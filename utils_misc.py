import re
from tkinter.filedialog import askopenfilename
from typing import Any

import openpyxl
import pandas as pd
from common import user_folder
from connectors import gcloud as gc

from date_utils import Event, EventName


def create_column_formatting(
    short_term_days: int = 14, long_term_days: int = 180
) -> dict[str, Any]:
    currency_columns = (
        f"avg sales dollar, {short_term_days} days",
        f"avg sales dollar, {long_term_days} days",
        "avg $",
        "avg price",
        "lost sales min",
        "lost sales max",
    )
    currency_formatting = {column: {"type": "currency"} for column in currency_columns}
    units_formatting = {
        column: {"type": "number"}
        for column in [
            "amz_inventory",
            "amz_available",
            "wh_inventory",
            "incoming_containers",
        ]
    }
    perc_formatting = {column: {"type": "percent"} for column in ["ISR", "ISR_short"]}
    column_formatting: dict[str, Any] = {
        "avg units": {
            "type": "3-color",
            "min_value": 2,
            "min_color": "red",
            "min_type": "num",
            "max_value": 10,
            "max_color": "green",
            "max_type": "num",
            "mid_value": 5,
            "mid_color": "yellow",
            "mid_type": "num",
        },
        "dos_available": [
            {
                "type": "3-color",
                "min_value": 21,
                "min_color": "red",
                "min_type": "num",
                "max_value": 49,
                "max_color": "green",
                "max_type": "num",
                "mid_value": 30,
                "mid_color": "yellow",
                "mid_type": "num",
            },
            {"type": "decimal", "precision": 1},
        ],
        "dos_inbound": [
            {
                "type": "3-color",
                "min_value": 49,
                "min_color": "red",
                "min_type": "num",
                "max_value": 90,
                "max_color": "green",
                "max_type": "num",
                "mid_value": 60,
                "mid_color": "yellow",
                "mid_type": "num",
            },
            {"type": "decimal", "precision": 1},
        ],
        "dos_shipped": [
            {
                "type": "3-color",
                "min_value": 49,
                "min_color": "red",
                "min_type": "num",
                "max_value": 90,
                "max_color": "green",
                "max_type": "num",
                "mid_value": 60,
                "mid_color": "yellow",
                "mid_type": "num",
            },
            {"type": "decimal", "precision": 1},
        ],
    }
    column_formatting.update(currency_formatting)
    column_formatting.update(units_formatting)
    column_formatting.update(perc_formatting)
    return column_formatting


def load_excel_with_hyperlinks(file_path, sheet_name: str | None = None):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active if not sheet_name else wb[sheet_name]
    if not sheet:
        raise ValueError("The Excel file does not contain any sheets.")

    data = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2):
        row_data = []
        for cell in row:
            val = cell.value

            if isinstance(val, str) and val.startswith("=HYPERLINK"):
                match = re.search(r',"(.*?)"\)', val)
                if match:
                    row_data.append(match.group(1))
                else:
                    # Fallback if regex fails (e.g. if formula structure differs)
                    row_data.append(val)
            else:
                row_data.append(val)
        data.append(row_data)

    return pd.DataFrame(data, columns=headers)


def push_restock_to_bq() -> None:
    """
    Pushes inventory restock to BigQuery table daily_reports.restock
    """
    file_path = askopenfilename(
        title="Select a file with the forecast", initialdir=user_folder
    )
    restock = load_excel_with_hyperlinks(file_path, sheet_name="restock")

    if (
        not isinstance(restock, pd.DataFrame)
        or restock.empty
        or "to_ship_units" not in restock.columns
    ):
        raise BaseException(
            "restock must be a non-empty DataFrame with 'to_ship_units' column"
        )
    _ = gc.push_to_cloud(
        restock, destination="daily_reports.restock", if_exists="replace"
    )


def push_forecast_to_bq() -> None:
    """
    Helper function to push forecast located in https://drive.google.com/drive/folders/1fSNHjoA6o1EOLOuBZrIrKDcM3wG9Xyre?usp=drive_link
    to BigQuery table daily_reports.forecast
    """

    file_path = askopenfilename(
        title="Select a file with the forecast", initialdir=user_folder
    )
    forecast = pd.read_excel(file_path)

    if (
        not isinstance(forecast, pd.DataFrame)
        or forecast.empty
        or "units" not in forecast.columns
    ):
        raise BaseException(
            "forecast must be a non-empty DataFrame with 'to_ship_units' column"
        )
    forecast = forecast[["asin", "date", "units", "$"]]
    _ = gc.push_to_cloud(
        forecast, destination="daily_reports.forecast", if_exists="replace"
    )


def compare_events():
    """Helper function to compare hourly sales per ASIN between events"""
    pd2025 = Event(name=EventName.PD, year=2025, month=7, start=8, duration=4)
    pd2026 = Event(name=EventName.PD, year=2026, month=6, start=23, duration=4)

    def _generate_query(event: Event):
        return f"""
        SELECT
            DATETIME(purchase_date, "America/Los_Angeles") as pacific_datetime, 
            sales_channel,
            asin,
            quantity,
            item_price,
            item_promotion_discount,
            order_status, item_status
            from `mellanni-project-da.reports.all_orders`
        WHERE DATETIME(purchase_date, "America/Los_Angeles") BETWEEN DATETIME("{event.event_start_str}") AND DATETIME("{event.event_end_str}")
        AND LOWER(sales_channel) = "amazon.com"
    """

    def _generate_hourly_template(asin_df):
        asins = asin_df["asin"].values.tolist()
        days = ["day 1", "day 2", "day 3", "day 4"]
        hours = range(24)

        columns = ["Day", "Hour", "ASIN"]
        full_template = pd.DataFrame(columns=columns)

        for day in days:
            for hour in hours:
                temp = pd.DataFrame(columns=columns)
                temp["ASIN"] = asins
                temp["Day"] = day
                temp["Hour"] = hour

                full_template = pd.concat([full_template, temp])

        return full_template

    query_event1 = _generate_query(pd2025)
    query_event2 = _generate_query(pd2026)

    with gc.gcloud_connect() as client:
        event1_result = client.query(query_event1).to_dataframe()
        event2_result = client.query(query_event2).to_dataframe()
        dict_result = client.query(
            "select distinct(asin) from  `mellanni-project-da.auxillary_development.dictionary`"
        ).to_dataframe()

    event1_result["pacific_datetime"] = event1_result["purchase_date"].dt.tz_localize(
        None
    )
    event2_result["pacific_datetime"] = event2_result["purchase_date"].dt.tz_localize(
        None
    )
    asin_template = _generate_hourly_template(dict_result)

    with pd.ExcelWriter(
        path="/home/misunderstood/temp/results.xlsx", engine="xlsxwriter"
    ) as writer:
        event1_result.to_excel(writer, sheet_name="pd2025", index=False)
        event2_result.to_excel(writer, sheet_name="pd2026", index=False)
        asin_template.to_excel(writer, sheet_name="full template", index=False)
